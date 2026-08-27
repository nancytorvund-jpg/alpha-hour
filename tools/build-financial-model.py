#!/usr/bin/env python3
"""Build Alpha-Hours-Financial-Model.xlsx (written to OUT, outside the repo).

Staggered rollout (Aug 27, 2026 revision): Session 2 New York only; Session 3
adds Greenwich and both Boston campuses; Session 4 adds Chicago (517) and the
three Florida campuses; Session 5 holds; national launch Session 1, Fall 2027.

Every cross-reference is derived from a key -> row dictionary that is filled
while rows are placed. No cell address is hand-typed anywhere in a formula.
Only formulas that behave identically in Excel and Google Sheets are used:
SUM, ROUND, ROUNDUP, IF and plain arithmetic. Per-campus-per-session values
are computed in explicit helper rows and summed (no array tricks).

Run:  python3 tools/build-financial-model.py
Import: build() returns (workbook, keys) where keys maps a name to a
        "'Sheet'!$C$R" reference for verification.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT = "/private/tmp/claude-501/-Users-nancytorvund/feceb8c1-fb3f-4825-aa0e-4efcfbaa183d/scratchpad/Alpha-Hours-Financial-Model.xlsx"

INK = "12100C"
HDR_FONT = Font(bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor=INK)
SEC_FONT = Font(bold=True, size=13, color=INK)
NOTE_FONT = Font(italic=True, color="6B6459")
TITLE_FONT = Font(bold=True, size=16, color=INK)
INPUT_FILL = PatternFill("solid", fgColor="FFF4DA")
TOTAL_FONT = Font(bold=True)
USD = '"$"#,##0'
PCT = '0%'
PCT1 = '0.0%'
NUM = '#,##0'
TEXT = '@'


def col_letter(col):
    return col if isinstance(col, str) else get_column_letter(col)


class Tab:
    """A worksheet plus a key -> row map built while rows are placed."""

    def __init__(self, wb, name, first=False):
        if first:
            self.ws = wb.active
            self.ws.title = name
        else:
            self.ws = wb.create_sheet(name)
        self.name = name
        self.rows = {}
        self.r = 4  # first row under the title block

    # -- placement --------------------------------------------------------
    def title(self, text, sub=None):
        self.ws["A1"] = text
        self.ws["A1"].font = TITLE_FONT
        if sub:
            self.ws["A2"] = sub
            self.ws["A2"].font = NOTE_FONT

    def header(self, cols, row=None):
        row = self.r if row is None else row
        for i, c in enumerate(cols, 1):
            cell = self.ws.cell(row=row, column=i, value=c)
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        if row == self.r:
            self.r += 1
        return row

    def section(self, text):
        self.ws.cell(row=self.r, column=1, value=text).font = SEC_FONT
        self.r += 1

    def blank(self, n=1):
        self.r += n

    def put(self, row, col, value, fmt=None, inp=False, bold=False, italic=False):
        cell = self.ws.cell(row=row, column=col if isinstance(col, int) else
                            self.ws[f"{col}1"].column, value=value)
        if fmt:
            cell.number_format = fmt
        if inp:
            cell.fill = INPUT_FILL
        if bold:
            cell.font = TOTAL_FONT
        if italic:
            cell.font = NOTE_FONT
        return cell

    def line(self, key, label, value, fmt=None, inp=None, bold=False, col=2):
        """Place a label in column A and a value in `col`, register `key`."""
        row = self.r
        self.ws.cell(row=row, column=1, value=label)
        if bold:
            self.ws.cell(row=row, column=1).font = TOTAL_FONT
        if value is not None:
            is_formula = isinstance(value, str) and value.startswith("=")
            self.put(row, col, value, fmt, inp=(not is_formula) if inp is None else inp, bold=bold)
        if key:
            assert key not in self.rows, f"duplicate key {key} on {self.name}"
            self.rows[key] = row
        self.r += 1
        return row

    def register(self, key, row):
        assert key not in self.rows, f"duplicate key {key} on {self.name}"
        self.rows[key] = row
        return row

    # -- references -------------------------------------------------------
    def row(self, key):
        return self.rows[key]

    def ref(self, key, col="B"):
        """Absolute reference usable from any sheet."""
        return f"'{self.name}'!${col_letter(col)}${self.rows[key]}"

    def at(self, key, col="B"):
        """Same-sheet relative reference."""
        return f"{col_letter(col)}{self.rows[key]}"

    def widths(self, w):
        for i, x in enumerate(w, 1):
            self.ws.column_dimensions[get_column_letter(i)].width = x

    def fmt_range(self, r1, r2, colfmts):
        for r in range(r1, r2 + 1):
            for col, f in colfmts:
                self.ws[f"{col}{r}"].number_format = f


def is_formula(v):
    return isinstance(v, str) and v.startswith("=")


def build():
    wb = Workbook()
    keys = {}  # name -> absolute reference, for verification

    # ================================================================== #
    # Assumptions
    # ================================================================== #
    A = Tab(wb, "Assumptions")
    A.title("Alpha Hours financial model: assumptions",
            "Yellow cells are inputs. Every other tab is formulas off this one. "
            "Session-based plan, staggered rollout, August 27, 2026.")
    a = A.at  # same-sheet reference helper

    A.section("PRICING")
    A.line("price", "Price per session, premium markets", 4500, USD)
    A.line("price_std", "Price per session, standard markets", 3500, USD)
    A.line("sess", "Sessions per year", 5, NUM)
    A.line("coh", "Cohorts per seat (enrolled = seats x this)", 4, NUM)
    A.line("disc", "Current Alpha family discount", 0.05, PCT)
    A.line("cont", "Continuation price from sixth session, premium", 5000, USD)
    A.line("deposit", "Deposit at sign-up", 500, USD)
    A.line("break_price", "Break week price, per week", 4500, USD)

    A.section("VARIABLE COST PER CHILD PER SESSION")
    A.line("tb", "Timeback license", 150, USD)
    A.line("mot", "Extrinsic motivation model (Alpha bucks, Emporium at cost)", 40, USD)
    A.line("sup", "Supplies", 15, USD)
    A.line("snk", "Snacks (about $2 a weekday block x14, $4 a Saturday block x7)", 30, USD)
    A.line("shw", "AlphaTest Showcase, per child (printed mastery map, certificate)", 5, USD)
    A.line("proc_rate", "Payment processing rate", 0.03, PCT)
    A.line("proc", "Payment processing per child per session (rate x price)",
           f"=ROUND({a('price')}*{a('proc_rate')},0)", USD)
    A.line("mkt", "Local marketing", 60, USD)
    A.line("var_s", "Variable cost per child per session",
           f"={a('tb')}+{a('mot')}+{a('sup')}+{a('snk')}+{a('shw')}+{a('proc')}+{a('mkt')}", USD, bold=True)
    A.line("var_y", "Variable cost per child per year", f"={a('var_s')}*{a('sess')}", USD, bold=True)

    A.section("GUIDES")
    A.line("guide", "Guide salary, part-time, per year", 50000, USD)
    A.line("ratio", "Seats per guide (one guide per 10 seats)", 10, NUM)
    A.line("guide_child", "Guide cost per enrolled child per year at even fill",
           f"={a('guide')}/({a('ratio')}*{a('coh')})", USD)
    A.line("min_guides", "Minimum adults per open campus (guides)", 2, NUM)

    A.section("FIXED COST PER CAMPUS PER YEAR")
    A.line("lg", "Lead Guide", 75000, USD)
    A.line("coord", "Campus Coordinator (50 percent role)", 50000, USD)
    A.line("fac", "Facility (custodial, HVAC, security after hours)", 45000, USD)
    A.line("ins", "Insurance", 8000, USD)
    A.line("shw_ev", "AlphaTest Showcase events, per campus per session", 2000, USD)
    A.line("fixed", "Fixed cost per campus per year",
           f"={a('lg')}+{a('coord')}+{a('fac')}+{a('ins')}+{a('shw_ev')}*{a('sess')}", USD, bold=True)
    A.line("mincost", "Minimum operating cost (fixed plus minimum guides)",
           f"={a('fixed')}+{a('min_guides')}*{a('guide')}", USD)
    A.line("contrib", "Contribution per child per year before guides",
           f"=({a('price')}-{a('var_s')})*{a('sess')}", USD)
    A.line("breakeven", "Break-even enrolled students per campus",
           f"=ROUNDUP({a('mincost')}/{a('contrib')},0)", NUM, bold=True)
    A.line("breakeven_cohort", "Break-even per cohort (about)",
           f"=ROUNDUP({a('breakeven')}/{a('coh')},0)", NUM)

    A.section("CENTRAL TEAM")
    A.line("central", "Central team, fully loaded, Year 1 (charged in full)", 875000, USD)
    A.line("central_nat", "Central team at full network", 4000000, USD)

    A.section("FILL RAMP BY EACH CAMPUS'S OWN SESSION COUNT, PLAN CASE")
    A.line("y1_first", "Year 1 opens in session number (Session 2, Oct 19 to Dec 18, 2026)", 2, NUM)
    A.line("ramp1", "A campus's first session open", 0.6, PCT)
    A.line("ramp2", "A campus's second session", 0.75, PCT)
    A.line("ramp3", "A campus's third session", 0.9, PCT)
    A.line("ramp4", "A campus's fourth session and later", 1.0, PCT)
    A.line("fcap", "Capacity case fill, every open campus, every session", 1.0, PCT)
    A.line("ffloor", "Floor case fill, every open campus, every session", 0.25, PCT)

    A.section("SENSITIVITY LEVELS")
    for i, f in enumerate((0.25, 0.5, 0.75, 1.0), 1):
        A.line(f"sf{i}", f"Fill level {i}", f, PCT)
    for i, p in enumerate((3500, 4000, 4500, 5000, 5500), 1):
        A.line(f"sp{i}", f"Session price level {i}", p, USD)
    A.line("price_step", "Price step for the per-step revenue line", 500, USD)

    A.section("LANE MIX AND VALUE")
    A.line("stay", "Stay (never enrolling)", 0.70, PCT)
    A.line("cons", "Considering", 0.15, PCT)
    A.line("pty", "Path to yes", 0.05, PCT)
    A.line("cur", "Current Alpha families", 0.10, PCT)
    A.line("stay_years", "Stay lane modeled tenure, years", 3, NUM)
    A.line("cons_sess", "Considering lane sessions before deciding", 2, NUM)
    A.line("pty_sess", "Path to yes lane sessions before deciding", 3, NUM)
    A.line("lane_conv", "Modeled conversion (Considering) / admit rate (Path to yes)", 0.20, PCT)
    A.line("ltv", "Lifetime tuition per conversion, midpoint", 240000, USD)
    for i, c in enumerate((0.04, 0.08, 0.12), 1):
        A.line(f"fd{i}", f"Feeder conversion rate, case {i}", c, PCT)

    A.section("NATIONAL THREE-YEAR VIEW")
    A.line("eve_ceiling", "Evening block ceiling as a share of K-8 capacity revenue", 0.5, PCT)
    A.line("break_weeks", "Break weeks per year from Year 2", 2, NUM)
    A.line("y2_k8", "Year 2 K-8 fill", 0.75, PCT)
    A.line("y2_eve", "Year 2 evening block, share of its ceiling", 0.40, PCT)
    A.line("y2_brk", "Year 2 break weeks, share of capacity", 0.60, PCT)
    A.line("y3_k8", "Year 3 K-8 fill", 0.90, PCT)
    A.line("y3_eve", "Year 3 evening block, share of its ceiling", 0.70, PCT)
    A.line("y3_brk", "Year 3 break weeks, share of capacity", 0.80, PCT)
    A.widths([70, 16])

    for k in ("price", "var_s", "var_y", "fixed", "mincost", "contrib", "breakeven", "proc"):
        keys["A." + k] = A.ref(k)

    # ================================================================== #
    # Campuses
    # ================================================================== #
    C = Tab(wb, "Campuses")
    C.title("Pilot campuses at capacity, annual",
            "Opens in session (yellow) drives the Year 1 tab. Enrolled = seats x cohorts. "
            "Guides = seats / seats per guide, rounded up. William Street shares New York's coordinator.")
    C.header(["Campus", "Seats", "Opens in session", "Shares coordinator (1 = yes)", "Enrolled",
              "Revenue / year", "Variable cost", "Guides", "Guide cost", "Fixed cost / year",
              "Site profit", "Margin", "Campuses"])
    camps = [
        ("maiden", "New York, 180 Maiden Lane", 100, 2, 0),
        ("william", "New York, 156 William Street", 30, 2, 1),
        ("greenwich", "Greenwich (Armonk)", 80, 3, 0),
        ("bostonA", "Boston A", 40, 3, 0),
        ("bostonB", "Boston B", 100, 3, 0),
        ("chicago", "Chicago", 517, 4, 0),
        ("miami", "Miami", 150, 4, 0),
        ("palm", "Palm Beach", 60, 4, 0),
        ("miamibeach", "Miami Beach", 50, 4, 0),
    ]
    camp_keys = [k for k, *_ in camps]
    launch_keys = camp_keys[:2]
    s3_keys = camp_keys[:5]
    full_keys = camp_keys
    for key, name, seats, opens, shares in camps:
        r = C.r
        C.put(r, 1, name)
        C.put(r, 2, seats, NUM, inp=True)
        C.put(r, 3, opens, NUM, inp=True)
        C.put(r, 4, shares, NUM, inp=True)
        C.put(r, 5, f"=B{r}*{A.ref('coh')}", NUM)
        C.put(r, 6, f"=E{r}*{A.ref('price')}*{A.ref('sess')}", USD)
        C.put(r, 7, f"=E{r}*{A.ref('var_y')}", USD)
        C.put(r, 8, f"=ROUNDUP(B{r}/{A.ref('ratio')},0)", NUM)
        C.put(r, 9, f"=H{r}*{A.ref('guide')}", USD)
        C.put(r, 10, f"={A.ref('fixed')}-D{r}*{A.ref('coord')}", USD)
        C.put(r, 11, f"=F{r}-G{r}-I{r}-J{r}", USD)
        C.put(r, 12, f"=K{r}/F{r}", PCT)
        C.put(r, 13, f"=IF(B{r}>0,1,0)", NUM)
        C.register(key, r)
        C.r += 1

    def plus(tab, ks, col):
        return "+".join(tab.at(k, col) for k in ks)

    def total_row(key, label, ks):
        r = C.r
        C.put(r, 1, label, bold=True)
        for col, fmt in (("B", NUM), ("E", NUM), ("F", USD), ("G", USD), ("H", NUM),
                         ("I", USD), ("J", USD), ("K", USD), ("M", NUM)):
            C.put(r, C.ws[f"{col}1"].column, f"={plus(C, ks, col)}", fmt, bold=True)
        C.put(r, 12, f"=K{r}/F{r}", PCT, bold=True)
        C.register(key, r)
        C.r += 1

    total_row("launch", "Launch total, Session 2 (New York)", launch_keys)
    total_row("s3", "Session 3 total (plus Greenwich, Boston)", s3_keys)
    total_row("full", "Full pilot total, Sessions 4 and 5", full_keys)
    C.blank()
    for fk, lab in (("launch", "Per session at capacity, launch footprint (New York)"),
                    ("s3", "Per session at capacity, Session 3 footprint"),
                    ("full", "Per session at capacity, full pilot footprint")):
        r = C.line(f"per_sess_{fk}", lab, None)
        C.put(r, 6, f"={C.at(fk, 'F')}/{A.ref('sess')}", USD)
        C.put(r, 11, f"={C.at(fk, 'K')}/{A.ref('sess')}", USD)
    r = C.line("full_net", "Full pilot run rate, net after central team", None)
    C.put(r, 11, f"={C.at('full', 'K')}-{A.ref('central')}", USD, bold=True)
    C.widths([44, 8, 10, 14, 10, 15, 14, 8, 13, 13, 14, 9, 10])

    for k in ("launch", "s3", "full"):
        for col in "BEFGHIJKM":
            keys[f"C.{k}.{col}"] = C.ref(k, col)
    keys["C.full_net.K"] = C.ref("full_net", "K")

    # ================================================================== #
    # Year 1
    # ================================================================== #
    Y = Tab(wb, "Year 1")
    Y.title("Year 1 by session: plan, capacity, floor",
            "A campus is open in a session once that session reaches its opening session (Campuses tab). "
            "Plan fill follows each campus's own session count (Assumptions ramp). Guides are hired to sold "
            "seats and fixed cost is charged only for open campuses; both are built campus by campus in the "
            "helper block under each case and summed.")
    dates = ["Oct 19 to Dec 18, 2026", "Jan 4 to Feb 19, 2027",
             "Feb 24 to Apr 16, 2027", "Apr 26 to Jun 18, 2027"]
    y1_cases = []

    def ramp_fill(r):
        k = f"B{r}-C{r}"
        return (f"=IF(E{r}=1,IF({k}=0,{A.ref('ramp1')},IF({k}=1,{A.ref('ramp2')},"
                f"IF({k}=2,{A.ref('ramp3')},{A.ref('ramp4')}))),0)")

    def flat_fill(akey):
        return lambda r: f"=IF(E{r}=1,{A.ref(akey)},0)"

    def year_block(case, label, fill_formula):
        Y.section(label)
        Y.header(["Session", "Dates", "Session number", "Campuses open", "Enrolled", "Revenue",
                  "Variable cost", "Guides", "Guide cost", "Fixed cost", "Site profit", "Margin"])
        first = Y.r
        sess_rows = []
        for i in range(4):
            r = Y.r
            Y.put(r, 1, f"Session {i + 2}")
            Y.put(r, 2, dates[i])
            Y.put(r, 3, f"={A.ref('y1_first')}" + (f"+{i}" if i else ""), NUM)
            # D, E, H, J are filled from the helper block below once it is placed
            Y.put(r, 6, f"=E{r}*{A.ref('price')}", USD)
            Y.put(r, 7, f"=E{r}*{A.ref('var_s')}", USD)
            Y.put(r, 9, f"=H{r}*{A.ref('guide')}/{A.ref('sess')}", USD)
            Y.put(r, 11, f"=F{r}-G{r}-I{r}-J{r}", USD)
            Y.put(r, 12, f"=K{r}/F{r}", PCT)
            Y.register(f"{case}_s{i + 2}", r)
            sess_rows.append(r)
            Y.r += 1
        last = sess_rows[-1]
        t = Y.r
        Y.put(t, 1, "Year 1 total", bold=True)
        for col, fmt in (("E", NUM), ("F", USD), ("G", USD), ("I", USD), ("J", USD), ("K", USD)):
            Y.put(t, Y.ws[f"{col}1"].column, f"=SUM({col}{first}:{col}{last})", fmt, bold=True)
        Y.put(t, 12, f"=K{t}/F{t}", PCT, bold=True)
        Y.register(f"{case}_total", t)
        Y.r += 1
        c = Y.line(f"{case}_central", "Central team (charged in full against Year 1)", None)
        Y.put(c, 11, f"=-{A.ref('central')}", USD)
        n = Y.line(f"{case}_net", "Net", None, bold=True)
        Y.put(n, 11, f"=K{t}+K{c}", USD, bold=True)

        # helper block: one row per campus per session
        Y.blank()
        Y.put(Y.r, 1, f"Campus by session, {label.split(' (')[0].lower()}: open flag, fill, "
                      "enrolled, guides hired to sold seats, fixed cost for open campuses", italic=True)
        Y.r += 1
        Y.header(["Campus", "Session number", "Opens in session", "Seats", "Open (1 = yes)", "Fill",
                  "Enrolled", "Guides", "Fixed cost, per session"])
        for j, sr in enumerate(sess_rows):
            hfirst = Y.r
            for ck in camp_keys:
                r = Y.r
                Y.put(r, 1, f"={C.ref(ck, 'A')}")
                Y.put(r, 2, f"=$C${sr}", NUM)
                Y.put(r, 3, f"={C.ref(ck, 'C')}", NUM)
                Y.put(r, 4, f"={C.ref(ck, 'B')}", NUM)
                Y.put(r, 5, f"=IF(B{r}>=C{r},1,0)", NUM)
                Y.put(r, 6, fill_formula(r), PCT)
                Y.put(r, 7, f"=ROUND(D{r}*{A.ref('coh')}*F{r},0)", NUM)
                Y.put(r, 8, f"=ROUNDUP(D{r}*F{r}/{A.ref('ratio')},0)", NUM)
                Y.put(r, 9, f"=IF(E{r}=1,{C.ref(ck, 'J')}/{A.ref('sess')},0)", USD)
                Y.r += 1
            hlast = Y.r - 1
            hs = Y.r
            Y.put(hs, 1, f"Session {j + 2} total", bold=True)
            for col, fmt in (("E", NUM), ("G", NUM), ("H", NUM), ("I", USD)):
                Y.put(hs, Y.ws[f"{col}1"].column, f"=SUM({col}{hfirst}:{col}{hlast})", fmt, bold=True)
            Y.register(f"{case}_h{j + 2}", hs)
            Y.put(sr, 4, f"=E{hs}", NUM)
            Y.put(sr, 5, f"=G{hs}", NUM)
            Y.put(sr, 8, f"=H{hs}", NUM)
            Y.put(sr, 10, f"=I{hs}", USD)
            Y.r += 1
        Y.r += 1
        y1_cases.append(case)

    year_block("plan", "Plan (60 / 75 / 90 / 100 percent by each campus's own session count, "
                       "the number the Head is held to)", ramp_fill)
    year_block("cap", "Capacity (every seat sold at every open campus)", flat_fill("fcap"))
    year_block("floor", "Floor (25 percent fill at every open campus all year)", flat_fill("ffloor"))

    Y.section("Three cases")
    Y.header(["Case", "", "", "", "", "Revenue", "", "", "", "", "Site profit", "Net after central team"])
    for case, lab in (("cap", "Capacity (100 percent at every open campus)"),
                      ("plan", "Plan (60 / 75 / 90 / 100 by each campus's own session count)"),
                      ("floor", "Floor (25 percent all year)")):
        r = Y.line(f"sum_{case}", lab, None)
        Y.put(r, 6, f"={Y.at(f'{case}_total', 'F')}", USD)
        Y.put(r, 11, f"={Y.at(f'{case}_total', 'K')}", USD)
        Y.put(r, 12, f"={Y.at(f'{case}_net', 'K')}", USD)
    Y.widths([44, 24, 10, 10, 10, 14, 13, 10, 12, 14, 14, 14])

    for case in y1_cases:
        for s in (2, 3, 4, 5):
            for col in "DEFHJK":
                keys[f"Y.{case}_s{s}.{col}"] = Y.ref(f"{case}_s{s}", col)
        keys[f"Y.{case}_total.F"] = Y.ref(f"{case}_total", "F")
        keys[f"Y.{case}_total.K"] = Y.ref(f"{case}_total", "K")
        keys[f"Y.{case}_net.K"] = Y.ref(f"{case}_net", "K")

    # ================================================================== #
    # 100-seat campus
    # ================================================================== #
    P = Tab(wb, "100-seat campus")
    P.title("A 100-seat campus at capacity, line by line, per year",
            "Change the seats (yellow) to see any campus.")
    P.line("seats", "Seats", 100, NUM)
    P.line("enr", "Enrolled", f"={P.at('seats')}*{A.ref('coh')}", NUM)
    P.blank()
    P.line("rev", "Revenue", f"={P.at('enr')}*{A.ref('price')}*{A.ref('sess')}", USD, bold=True)
    cost_first = P.r
    e = P.at("enr")
    S = A.ref("sess")
    P.line("c_tb", "Timeback", f"=-{e}*{A.ref('tb')}*{S}", USD)
    P.line("c_mot", "Extrinsic motivation model (Emporium)", f"=-{e}*{A.ref('mot')}*{S}", USD)
    P.line("c_sup", "Supplies", f"=-{e}*{A.ref('sup')}*{S}", USD)
    P.line("c_snk", "Snacks", f"=-{e}*{A.ref('snk')}*{S}", USD)
    P.line("c_shw", "AlphaTest Showcase (per child reports plus campus events)",
           f"=-({e}*{A.ref('shw')}*{S}+{A.ref('shw_ev')}*{S})", USD)
    P.line("c_proc", "Payment processing", f"=-{e}*{A.ref('proc')}*{S}", USD)
    P.line("c_mkt", "Local marketing", f"=-{e}*{A.ref('mkt')}*{S}", USD)
    P.line("c_guides", "Guides (seats / seats per guide, rounded up, at guide salary)",
           f"=-ROUNDUP({P.at('seats')}/{A.ref('ratio')},0)*{A.ref('guide')}", USD)
    P.line("c_lg", "Lead Guide", f"=-{A.ref('lg')}", USD)
    P.line("c_coord", "Campus Coordinator", f"=-{A.ref('coord')}", USD)
    P.line("c_fac", "Facility", f"=-{A.ref('fac')}", USD)
    P.line("c_ins", "Insurance", f"=-{A.ref('ins')}", USD)
    cost_last = P.r - 1
    P.blank()
    P.line("cost", "Total cost", f"=-SUM(B{cost_first}:B{cost_last})", USD, bold=True)
    P.line("profit", "Site profit", f"={P.at('rev')}-{P.at('cost')}", USD, bold=True)
    P.line("margin", "Margin", f"={P.at('profit')}/{P.at('rev')}", PCT1)
    P.blank()
    P.line("engage", "Motivation model, snacks and showcase together",
           f"=-({P.at('c_mot')}+{P.at('c_snk')}+{P.at('c_shw')})", USD)
    P.line("engage_share", "As a share of revenue", f"={P.at('engage')}/{P.at('rev')}", PCT1)
    P.widths([60, 16])
    for k in ("rev", "cost", "profit"):
        keys["P." + k] = P.ref(k)

    # ================================================================== #
    # Sensitivity
    # ================================================================== #
    Sn = Tab(wb, "Sensitivity")
    Sn.title("Sensitivity on the full pilot footprint (all nine campuses), annual",
             "Fill levels and price levels are inputs on the Assumptions tab.")
    Sn.section("Fill (guides hired to sold seats; per-campus counts in the helper block below)")
    Sn.header(["Fill", "Enrolled", "Revenue", "Variable cost", "Guides", "Guide cost",
               "Fixed cost", "Site profit", "Net after central", "Margin"])
    fill_rows = []
    for i in range(1, 5):
        r = Sn.r
        Sn.put(r, 1, f"={A.ref(f'sf{i}')}", PCT)
        Sn.put(r, 2, f"=ROUND({C.ref('full', 'B')}*{A.ref('coh')}*A{r},0)", NUM)
        Sn.put(r, 3, f"=B{r}*{A.ref('price')}*{A.ref('sess')}", USD)
        Sn.put(r, 4, f"=B{r}*{A.ref('var_y')}", USD)
        # E (guides) filled from helper block
        Sn.put(r, 6, f"=E{r}*{A.ref('guide')}", USD)
        Sn.put(r, 7, f"={C.ref('full', 'J')}", USD)
        Sn.put(r, 8, f"=C{r}-D{r}-F{r}-G{r}", USD)
        Sn.put(r, 9, f"=H{r}-{A.ref('central')}", USD)
        Sn.put(r, 10, f"=H{r}/C{r}", PCT)
        Sn.register(f"fill{i}", r)
        fill_rows.append(r)
        Sn.r += 1
    Sn.blank()
    Sn.put(Sn.r, 1, "Guides by campus at each fill level (full pilot footprint)", italic=True)
    Sn.r += 1
    Sn.header(["Campus", "Seats", "", "", "Guides at fill 1", "Guides at fill 2",
               "Guides at fill 3", "Guides at fill 4"])
    hfirst = Sn.r
    for ck in full_keys:
        r = Sn.r
        Sn.put(r, 1, f"={C.ref(ck, 'A')}")
        Sn.put(r, 2, f"={C.ref(ck, 'B')}", NUM)
        for j, fr in enumerate(fill_rows):
            Sn.put(r, 5 + j, f"=ROUNDUP(B{r}*$A${fr}/{A.ref('ratio')},0)", NUM)
        Sn.r += 1
    hlast = Sn.r - 1
    hs = Sn.r
    Sn.put(hs, 1, "Guides on payroll", bold=True)
    for j, fr in enumerate(fill_rows):
        col = get_column_letter(5 + j)
        Sn.put(hs, 5 + j, f"=SUM({col}{hfirst}:{col}{hlast})", NUM, bold=True)
        Sn.put(fr, 5, f"={col}{hs}", NUM)
    Sn.r += 2

    Sn.section("Price, at capacity (payment processing scales with price)")
    Sn.header(["Session price", "Enrolled", "Revenue", "Variable cost", "Guides", "Guide cost",
               "Fixed cost", "Site profit", "Net after central", "Margin"])
    for i in range(1, 6):
        r = Sn.r
        Sn.put(r, 1, f"={A.ref(f'sp{i}')}", USD)
        Sn.put(r, 2, f"={C.ref('full', 'E')}", NUM)
        Sn.put(r, 3, f"=B{r}*A{r}*{A.ref('sess')}", USD)
        Sn.put(r, 4, f"=B{r}*({A.ref('var_s')}-{A.ref('proc')}+ROUND(A{r}*{A.ref('proc_rate')},0))*{A.ref('sess')}", USD)
        Sn.put(r, 5, f"={C.ref('full', 'H')}", NUM)
        Sn.put(r, 6, f"={C.ref('full', 'I')}", USD)
        Sn.put(r, 7, f"={C.ref('full', 'J')}", USD)
        Sn.put(r, 8, f"=C{r}-D{r}-F{r}-G{r}", USD)
        Sn.put(r, 9, f"=H{r}-{A.ref('central')}", USD)
        Sn.put(r, 10, f"=H{r}/C{r}", PCT)
        Sn.register(f"price{i}", r)
        Sn.r += 1
    Sn.blank()
    r = Sn.line("per_step", "Every price step (Assumptions) on the session price, at capacity, per year", None)
    Sn.put(r, 3, f"={C.ref('full', 'E')}*{A.ref('price_step')}*{A.ref('sess')}*(1-{A.ref('proc_rate')})", USD)
    Sn.widths([34, 10, 14, 13, 12, 12, 12, 14, 16, 9])
    for i in range(1, 5):
        keys[f"S.fill{i}.B"] = Sn.ref(f"fill{i}", "B")
        keys[f"S.fill{i}.C"] = Sn.ref(f"fill{i}", "C")
        keys[f"S.fill{i}.H"] = Sn.ref(f"fill{i}", "H")
    for i in range(1, 6):
        keys[f"S.price{i}.C"] = Sn.ref(f"price{i}", "C")
        keys[f"S.price{i}.H"] = Sn.ref(f"price{i}", "H")
    keys["S.per_step.C"] = Sn.ref("per_step", "C")

    # ================================================================== #
    # Lanes
    # ================================================================== #
    L = Tab(wb, "Lanes")
    L.title("Revenue by lane at full pilot capacity",
            "The plan is sized on Stay. Mix is an assumption Session 2 replaces.")
    L.header(["Lane", "Share", "Families", "Session revenue / year",
              "Value beyond the session fee, per family (modeled)", "Basis"])
    full_enr = C.ref("full", "E")
    pr, ss = A.ref("price"), A.ref("sess")
    lanes = [
        ("stay", "Stay (never enrolling, the business)", "stay",
         f"={pr}*{ss}+{A.ref('cont')}*{ss}*({A.ref('stay_years')}-1)",
         "Modeled tenure: year one at list, later years at the continuation price"),
        ("cons", "Considering (upside)", "cons",
         f"={pr}*{A.ref('cons_sess')}+{A.ref('lane_conv')}*{A.ref('ltv')}",
         "Sessions before deciding plus modeled conversion at the lifetime tuition midpoint"),
        ("pty", "Path to yes (upside)", "pty",
         f"={pr}*{A.ref('pty_sess')}+{A.ref('lane_conv')}*{A.ref('ltv')}",
         "Sessions before deciding plus modeled admit rate at the lifetime tuition midpoint"),
        ("cur", "Current Alpha families", "cur", None,
         "Retention of a family already paying full tuition; pays the current-family discount"),
    ]
    lfirst = L.r
    for key, name, share_key, val, basis in lanes:
        r = L.r
        L.put(r, 1, name)
        L.put(r, 2, f"={A.ref(share_key)}", PCT)
        L.put(r, 3, f"=ROUND({full_enr}*B{r},0)", NUM)
        if key == "cur":
            L.put(r, 4, f"=C{r}*{pr}*{ss}*(1-{A.ref('disc')})", USD)
        else:
            L.put(r, 4, f"=C{r}*{pr}*{ss}", USD)
        if val:
            L.put(r, 5, val, USD)
        L.put(r, 6, basis)
        L.register(key, r)
        L.r += 1
    llast = L.r - 1
    r = L.line("lane_total", "Total", None, bold=True)
    for col, fmt in (("B", PCT), ("C", NUM), ("D", USD)):
        L.put(r, L.ws[f"{col}1"].column, f"=SUM({col}{lfirst}:{col}{llast})", fmt, bold=True)
    L.blank()
    L.section("Feeder (upside): conversions to full Alpha")
    L.header(["Conversion rate", "Enrollments, full pilot", "Lifetime tuition, full pilot",
              "Enrollments, full network", "Lifetime tuition, full network"])
    feeder_rows = []
    for i in range(1, 4):
        r = L.r
        L.put(r, 1, f"={A.ref(f'fd{i}')}", PCT)
        L.put(r, 2, f"=ROUND({full_enr}*A{r},0)", NUM)
        L.put(r, 3, f"=B{r}*{A.ref('ltv')}", USD)
        # D and E (full network) are filled once the National tab exists
        L.register(f"fd{i}", r)
        feeder_rows.append(r)
        L.r += 1
    L.widths([40, 9, 22, 22, 30, 70])
    for k in ("stay", "cons", "pty", "cur"):
        for col in "CD":
            keys[f"L.{k}.{col}"] = L.ref(k, col)
    for k in ("stay", "cons", "pty"):
        keys[f"L.{k}.E"] = L.ref(k, "E")
    for i in (1, 2, 3):
        for col in "BC":
            keys[f"L.fd{i}.{col}"] = L.ref(f"fd{i}", col)

    # ================================================================== #
    # National
    # ================================================================== #
    N = Tab(wb, "National")
    N.title("The national build: capacity figures, K-8 blocks, premium price",
            "Four steps one session apart through the pilot, then the national launch with the 2027-28 "
            "school year. Each step is released by the week-3 checkpoint of the session before it.")
    N.section("Steps (pilot steps come from the Campuses tab subtotals; Hold and National are yellow inputs)")
    N.header(["Step", "Session", "Campuses or group", "Campuses added", "Seats added",
              "Enrolled added", "Run rate added", "Run rate"])
    sfirst = N.r
    steps = [
        ("launch", "Launch", "Session 2, October 19, 2026", "New York (2)",
         f"={C.ref('launch', 'M')}", f"={C.ref('launch', 'B')}"),
        ("step2", "Step 2", "Session 3, January 4, 2027", "Greenwich, Boston (2)",
         f"={C.ref('s3', 'M')}-{C.ref('launch', 'M')}", f"={C.ref('s3', 'B')}-{C.ref('launch', 'B')}"),
        ("step3", "Step 3", "Session 4, February 24, 2027", "Chicago (517), Miami, Palm Beach, Miami Beach",
         f"={C.ref('full', 'M')}-{C.ref('s3', 'M')}", f"={C.ref('full', 'B')}-{C.ref('s3', 'B')}"),
        ("hold", "Hold", "Session 5, April 26, 2027",
         "None: nine campuses run a full session with no new openings", 0, 0),
        ("national", "National", "Session 1, Fall 2027",
         "Austin (300) and every remaining campus, about 30 at about 50 seats each "
         "(nine pilot campuses plus these make the network)", 31, 1850),
    ]
    for key, step, sess_label, group, ncamp, nseats in steps:
        r = N.r
        N.put(r, 1, step)
        N.put(r, 2, sess_label)
        N.put(r, 3, group)
        N.put(r, 4, ncamp, NUM, inp=not is_formula(ncamp))
        N.put(r, 5, nseats, NUM, inp=not is_formula(nseats))
        N.put(r, 6, f"=E{r}*{A.ref('coh')}", NUM)
        N.put(r, 7, f"=F{r}*{A.ref('price')}*{A.ref('sess')}", USD)
        N.put(r, 8, f"=SUM($G${sfirst}:G{r})", USD)
        N.register(key, r)
        N.r += 1
    slast = N.r - 1
    N.blank()

    N.section("Full network at capacity")
    N.line("n_camp", "Campuses (nine pilot plus the national step)", f"=SUM(D{sfirst}:D{slast})", NUM)
    N.line("n_seats", "Seats", f"=SUM(E{sfirst}:E{slast})", NUM)
    N.line("n_enr", "Enrolled", f"=SUM(F{sfirst}:F{slast})", NUM)
    N.line("n_rev", "Revenue", f"={N.at('n_enr')}*{A.ref('price')}*{A.ref('sess')}", USD, bold=True)
    N.line("n_var", "Variable cost", f"=-{N.at('n_enr')}*{A.ref('var_y')}", USD)
    N.line("n_guides", "Guides", f"=ROUNDUP({N.at('n_seats')}/{A.ref('ratio')},0)", NUM)
    N.line("n_gc", "Guide cost", f"=-{N.at('n_guides')}*{A.ref('guide')}", USD)
    N.line("n_fx", "Fixed cost", f"=-{N.at('n_camp')}*{A.ref('fixed')}", USD)
    N.line("n_prof", "Site profit",
           f"={N.at('n_rev')}+{N.at('n_var')}+{N.at('n_gc')}+{N.at('n_fx')}", USD, bold=True)
    N.line("n_margin", "Site margin", f"={N.at('n_prof')}/{N.at('n_rev')}", PCT1)
    N.line("n_central", "Central team at full network", f"=-{A.ref('central_nat')}", USD)
    N.line("n_net", "Net", f"={N.at('n_prof')}+{N.at('n_central')}", USD, bold=True)
    N.line("n_pilot_margin", "Full pilot site margin (used for the three-year view)",
           f"={C.ref('full', 'L')}", PCT1)
    N.blank()

    N.section("Three-year view (network held at the campus count above; evening block and break weeks as additions)")
    N.header(["Year", "K-8 fill", "K-8 revenue", "Evening block share of its ceiling",
              "Evening block revenue", "Break weeks share", "Break week revenue",
              "Total revenue", "Net (about, at the full pilot site margin)"])
    for yk, ylabel in (("y2", "Year 2, SY27-28"), ("y3", "Year 3, SY28-29")):
        r = N.r
        N.put(r, 1, ylabel)
        N.put(r, 2, f"={A.ref(yk + '_k8')}", PCT)
        N.put(r, 3, f"={N.at('n_rev')}*B{r}", USD)
        N.put(r, 4, f"={A.ref(yk + '_eve')}", PCT)
        N.put(r, 5, f"={N.at('n_rev')}*{A.ref('eve_ceiling')}*D{r}", USD)
        N.put(r, 6, f"={A.ref(yk + '_brk')}", PCT)
        N.put(r, 7, f"={N.at('n_seats')}*{A.ref('break_weeks')}*{A.ref('break_price')}*F{r}", USD)
        N.put(r, 8, f"=C{r}+E{r}+G{r}", USD)
        N.put(r, 9, f"=H{r}*{N.at('n_pilot_margin')}-{A.ref('central_nat')}", USD)
        N.register(yk, r)
        N.r += 1
    N.widths([30, 28, 70, 12, 12, 14, 16, 16, 22])
    for k in ("n_camp", "n_seats", "n_enr", "n_rev", "n_prof", "n_net"):
        keys["N." + k] = N.ref(k)
    for sk in ("launch", "step2", "step3", "hold", "national"):
        keys[f"N.{sk}.H"] = N.ref(sk, "H")
    for yk in ("y2", "y3"):
        keys[f"N.{yk}.H"] = N.ref(yk, "H")
        keys[f"N.{yk}.I"] = N.ref(yk, "I")

    # full-network feeder columns on the Lanes tab (needs N.n_enr)
    for r in feeder_rows:
        L.put(r, 4, f"=ROUND({N.ref('n_enr')}*A{r},0)", NUM)
        L.put(r, 5, f"=D{r}*{A.ref('ltv')}", USD)
    keys["L.fd2.E"] = L.ref("fd2", "E")

    # ================================================================== #
    # Launch budget
    # ================================================================== #
    B = Tab(wb, "Launch budget")
    B.title("One-time launch budget, New York, before Session 2",
            "Yellow cells are inputs. Guide and seat counts come from the Campuses tab launch row. "
            "Each later step is sized the same way and paid from the sessions before it.")
    B.section("Basis inputs")
    B.line("lb_weeks", "Paid training weeks before launch", 2, NUM)
    B.line("lb_leads", "Lead Guides in the training cohort", 2, NUM)
    B.line("lb_wpy", "Weeks per year (part-time rate divisor)", 52, NUM)
    B.line("lb_markets", "Launch markets", 1, NUM)
    B.line("lb_mkt_per", "Launch marketing per market", 40000, USD)
    B.line("lb_dev_share", "Loaner device pool, share of launch seats", 0.2, PCT)
    B.line("lb_dev_cost", "Cost per loaner device", 600, USD)
    B.line("lb_emp", "Emporium opening inventory per seat", 25, USD)
    B.blank()
    B.header(["Item", "Basis", "Cost"])
    lfirst = B.r
    items = [
        ("Guide training cohort",
         "New York guides plus Lead Guides, paid training weeks at the part-time rate",
         f"=({C.ref('launch', 'H')}+{B.at('lb_leads')})*{A.ref('guide')}/{B.at('lb_wpy')}*{B.at('lb_weeks')}"),
        ("State exemption filing and legal", "New York", 15000),
        ("Launch marketing", "Per market x markets (New York)", f"={B.at('lb_mkt_per')}*{B.at('lb_markets')}"),
        ("Loaner device pool", "Share of launch seats x device cost",
         f"={C.ref('launch', 'B')}*{B.at('lb_dev_share')}*{B.at('lb_dev_cost')}"),
        ("Insurance deposits", "Two sites", 8000),
        ("Emporium opening inventory", "Per seat x launch seats", f"={C.ref('launch', 'B')}*{B.at('lb_emp')}"),
        ("Registration and payments stack", "Stripe, HubSpot, roster tooling", 30000),
    ]
    for name, basis, cost in items:
        r = B.r
        B.put(r, 1, name)
        B.put(r, 2, basis)
        B.put(r, 3, cost, USD, inp=not is_formula(cost))
        B.r += 1
    llast = B.r - 1
    r = B.line("lb_total", "Total", None, bold=True)
    B.put(r, 3, f"=SUM(C{lfirst}:C{llast})", USD, bold=True)
    B.blank()
    r = B.line("lb_deposits", "Session 2 deposits on plan enrollment", None)
    B.put(r, 3, f"={Y.ref('plan_s2', 'E')}*{A.ref('deposit')}", USD)
    r = B.line("lb_covered", "Deposits cover the launch budget", None)
    B.put(r, 3, f'=IF({B.at("lb_deposits", "C")}>={B.at("lb_total", "C")},"Yes","No")')
    B.widths([44, 70, 14])
    keys["B.lb_total.C"] = B.ref("lb_total", "C")
    keys["B.lb_deposits.C"] = B.ref("lb_deposits", "C")
    keys["B.lb_covered.C"] = B.ref("lb_covered", "C")

    # ================================================================== #
    # Summary (placed first)
    # ================================================================== #
    Sm = Tab(wb, "Summary", first=True)
    Sm.title("Alpha Hours financial model: summary",
             "Every figure is a formula off the Assumptions tab. Yellow cells are inputs.")
    Sm.header(["Headline", "Measure", "Figure"])
    summary = [
        ("Launch, Session 2 (New York only)", [
            ("Seats", f"={C.ref('launch', 'B')}", NUM),
            ("Enrolled", f"={C.ref('launch', 'E')}", NUM),
            ("Revenue run rate", f"={C.ref('launch', 'F')}", USD),
            ("Site profit", f"={C.ref('launch', 'K')}", USD)]),
        ("Session 3 footprint (New York, Greenwich, Boston)", [
            ("Seats", f"={C.ref('s3', 'B')}", NUM),
            ("Enrolled", f"={C.ref('s3', 'E')}", NUM),
            ("Revenue run rate", f"={C.ref('s3', 'F')}", USD),
            ("Site profit", f"={C.ref('s3', 'K')}", USD)]),
        ("Full pilot footprint, Sessions 4 and 5 (plus Chicago and Florida)", [
            ("Seats", f"={C.ref('full', 'B')}", NUM),
            ("Enrolled", f"={C.ref('full', 'E')}", NUM),
            ("Revenue run rate", f"={C.ref('full', 'F')}", USD),
            ("Site profit", f"={C.ref('full', 'K')}", USD),
            ("Net after central team", f"={C.ref('full_net', 'K')}", USD)]),
        ("Year 1 plan (Sessions 2 to 5, each campus at 60 / 75 / 90 / 100 percent by its own session count)", [
            ("Revenue", f"={Y.ref('plan_total', 'F')}", USD),
            ("Site profit", f"={Y.ref('plan_total', 'K')}", USD),
            ("Net after central team", f"={Y.ref('plan_net', 'K')}", USD)]),
        ("Year 1 capacity", [
            ("Revenue", f"={Y.ref('cap_total', 'F')}", USD),
            ("Net after central team", f"={Y.ref('cap_net', 'K')}", USD)]),
        ("Year 1 floor (25 percent fill all year)", [
            ("Revenue", f"={Y.ref('floor_total', 'F')}", USD),
            ("Net after central team", f"={Y.ref('floor_net', 'K')}", USD)]),
        ("One-time launch budget (New York)", [
            ("Total", f"={B.ref('lb_total', 'C')}", USD),
            ("Session 2 deposits on plan enrollment", f"={B.ref('lb_deposits', 'C')}", USD),
            ("Deposits cover the launch budget", f"={B.ref('lb_covered', 'C')}", TEXT)]),
        ("Break-even per campus", [
            ("Enrolled students", f"={A.ref('breakeven')}", NUM),
            ("Per cohort (about)", f"={A.ref('breakeven_cohort')}", NUM)]),
        ("Full network at capacity (Session 1, Fall 2027, K-8)", [
            ("Campuses", f"={N.ref('n_camp')}", NUM),
            ("Enrolled", f"={N.ref('n_enr')}", NUM),
            ("Revenue", f"={N.ref('n_rev')}", USD),
            ("Net after central team", f"={N.ref('n_net')}", USD)]),
        ("Year 3 (network, with evening block and break weeks)", [
            ("Revenue (about)", f"={N.ref('y3', 'H')}", USD),
            ("Net (about)", f"={N.ref('y3', 'I')}", USD)]),
        ("Revenue from families who will never enroll (Stay lane)", [
            ("Share of the mix", f"={A.ref('stay')}", PCT),
            ("Session revenue a year at full pilot capacity", f"={L.ref('stay', 'D')}", USD)]),
        ("Upside: feeder at the middle conversion case, full pilot base", [
            ("Conversion rate", f"={L.ref('fd2', 'A')}", PCT),
            ("Full-Alpha enrollments", f"={L.ref('fd2', 'B')}", NUM),
            ("Lifetime tuition (on top of the P&L)", f"={L.ref('fd2', 'C')}", USD)]),
    ]
    for head, measures in summary:
        for i, (m, f, fmt) in enumerate(measures):
            r = Sm.r
            if i == 0:
                Sm.put(r, 1, head, bold=True)
            Sm.put(r, 2, m)
            Sm.put(r, 3, f, fmt)
            Sm.r += 1
        Sm.r += 1
    Sm.widths([70, 42, 18])

    # order, freeze panes
    order = ["Summary", "Assumptions", "Campuses", "Year 1", "100-seat campus",
             "Sensitivity", "Lanes", "National", "Launch budget"]
    wb._sheets = [wb[n] for n in order]
    for ws in wb.worksheets:
        ws.freeze_panes = "A4"
    return wb, keys


if __name__ == "__main__":
    wb, keys = build()
    wb.save(OUT)
    print("saved", OUT)
    print("tabs", [ws.title for ws in wb.worksheets])
